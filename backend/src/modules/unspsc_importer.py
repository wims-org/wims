import argparse
import asyncio
import json
from datetime import datetime
from pathlib import Path

import openpyxl
import pymongo
from loguru import logger

from database_connector import MongoDBConnector
from models.database import Category

"""
This module handles the import of UNSPSC data from an Excel file into a MongoDB collection.
It checks if the data is already imported and complete by comparing MD5 hashes.
It is meant to be run as a standalone script or thread
"""


def dataclass_xlsx_reader(file_path: str, cls: type, expected_columns: list[str]):
    df = openpyxl.load_workbook(file_path)
    df1 = df.active
    columns = [df1.cell(row=1, column=col + 1).value for col in range(df1.max_column)]
    if not all(col in columns for col in expected_columns):
        missing_cols = [col for col in expected_columns if col not in columns]
        logger.error(f"column mismatch in Excel file: {missing_cols}")
        return

    for row in range(1, df1.max_row):
        row = [df1.cell(row=row + 1, column=col + 1).value for col in range(df1.max_column)]
        row_data = dict(zip(columns, row, strict=False))
        yield cls(**row_data)


def collection_is_complete(db_connector: MongoDBConnector, collection_name: str, filepath: str) -> bool:
    """Reads the hash file next to the data file and compares it to the current collection hash"""
    import_file = Path(filepath).parent / "unspsc_data.import.json"
    if not import_file.exists():
        logger.warning(f"Import file not found, unspsc import was never successful: {import_file}")
        return False
    result = json.loads(import_file.read_text() or "{}")
    expected_md5 = result.get("md5", {})
    if not expected_md5:
        logger.error(f"MD5 hash not found in import file: {import_file}")
        import_file.unlink(missing_ok=True)
        return False
    res_data = db_connector.db.command({"dbHash": 1, "collections": [collection_name]})
    md5 = res_data.get("collections", {}).get(collection_name, None)
    logger.debug(f"DB Hash for {collection_name}: {res_data}, {md5} expected: {expected_md5}")
    if md5 != expected_md5:
        logger.error(f"MD5 mismatch for {collection_name}: {md5} != {expected_md5}")
        import_file.unlink(missing_ok=True)
        return False
    return True


def write_md5_to_file(db_connector: MongoDBConnector, collection_name: str, filepath: str) -> None:
    """Puts a hash file next to the data file to indicate successful import"""
    import_file = Path(filepath).parent / "unspsc_data.import.json"
    res_data = db_connector.db.command({"dbHash": 1, "collections": [collection_name]})

    logger.debug(f"Writing DB Hash for {collection_name}: {res_data} to {import_file}")
    import_file.write_text(json.dumps({"md5": res_data.get("collections", {}).get(collection_name, None)}))


def import_unspsc_data(db_connector: MongoDBConnector, collection_name: str, file_path: str | None = None):
    # supported format https://www.ungm.org/Public/UNSPSC
    start_time = datetime.now()
    if file_path is None:
        file_path = str(Path(__file__).parent.parent.parent / "data" / "unspsc_data.xlsx")
    expected_columns = ["Key", "Parent key", "Code", "Title"]
    categories_batch = []

    # this could use bulk insert for better performance, wip
    for index, item in enumerate(
        dataclass_xlsx_reader(file_path=file_path, cls=Category, expected_columns=expected_columns)
    ):
        categories_batch.append(Category(**item.dict()).model_dump())
        if index % 1000 == 0:
            try:
                db_connector.db[collection_name].insert_many(categories_batch, ordered=False)
                logger.info(f"Importing UNSPSC data... processed {index} records")
                categories_batch = []
            except pymongo.errors.BulkWriteError as e:
                logger.warning(
                    f"Result importing UNSPSC batch: {[(k, len(v)) for k, v in e.details.items() if type(v) is list]}"
                )
    try:
        db_connector.db[collection_name].insert_many(categories_batch, ordered=False)  # remaining batch
    except pymongo.errors.BulkWriteError as e:
        logger.warning(
            f"Result importing UNSPSC batch: {[(k, len(v)) for k, v in e.details.items() if type(v) is list]}"
        )
    logger.info(f"UNSPSC data import completed successfully. Took {datetime.now() - start_time}")


def check_and_import_unspsc_data(db_connector: MongoDBConnector, collection_name: str, file_path: str | None = None):
    try:
        if collection_name not in db_connector.db.list_collection_names():
            db_connector.db.create_collection(collection_name)
            db_connector.db[collection_name].create_index("unspsc_code", unique=True)
            db_connector.db[collection_name].create_index("key", unique=True)
            db_connector.db[collection_name].create_index("parent_key")
            db_connector.db[collection_name].create_index("title")
            logger.info(f"Category collection '{collection_name}' created.")
        if not collection_is_complete(db_connector, collection_name, file_path or ""):
            import_unspsc_data(db_connector, collection_name, file_path)
            write_md5_to_file(db_connector, collection_name, file_path or "")
            logger.info(f"UNSPSC data imported into collection '{collection_name}'.")
        else:
            logger.info(f"UNSPSC data in collection '{collection_name}' is already complete.")
    except Exception as e:
        logger.exception(f"Error importing UNSPSC data: {e}")


if __name__ == "__main__":
    logger.info("Starting UNSPSC Importer")
    # read options from command line arguments for script usage
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--collection",
        type=str,
        default="categories",
        help="MongoDB collection name to import UNSPSC data into",
    )
    parser.add_argument(
        "--file-path",
        type=str,
        default=str(Path(__file__).parent.parent.parent / "data" / "unspsc_data.xlsx"),
        help="Path to the UNSPSC data Excel file",
    )
    parser.add_argument(
        "--config",
        type=dict,
        default={},
        help="Configuration dictionary",
    )
    args = parser.parse_args()

    db_connector = MongoDBConnector(
        uri=f"mongodb://{args.config.get('host', 'localhost')}:{args.config.get('port', '27017')}",
        database=args.config.get("database", "inventory"),
    )
    asyncio.run(
        check_and_import_unspsc_data(
            db_connector,
            collection_name=args.collection,
            file_path=args.file_path,
        )
    )
